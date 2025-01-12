from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth import authenticate,login,logout
from django.contrib import messages 
from django.contrib.auth.models import User
from user.models import CustomUser
from django.contrib.auth.decorators import user_passes_test
from django.views.decorators.cache import never_cache
from products.models import Category,Brand,Product,Variant
from django.contrib.auth import get_user_model
from orders.models import Order_items,Order
from django.db.models import Sum,F,Count
from django.utils import timezone
from django.db.models import Sum
from datetime import timedelta
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.db.models import Sum
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer
from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.utils import timezone
from datetime import datetime, timedelta
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth, TruncYear,TruncDate
from django.http import JsonResponse
import json
from dateutil import parser
import io

# Create your views here.

User=get_user_model()

#======================Admin login,logout=====================# 
@never_cache
def admin_login(request):
    if request.user.is_superuser:
        return redirect('dashboard')
    if request.method =='POST':
        username=request.POST.get('username')
        password=request.POST.get('password')
        user=authenticate(username=username,password=password)
        if user is not None:
            if user.is_superuser:
                login(request,user)
                return redirect('dashboard')
            else:
                messages.error(request,'You are not authorized user')
        else:
            messages.error(request,'invalid username or password!')
    return render(request,'admin/adminlogin.html')


@never_cache
@user_passes_test(lambda u : u.is_superuser,login_url='/adminn/')
def admin_logout(request):
    if request.method=='POST':
        logout(request)
    return redirect('admin_login')

#======================Admin login,logout End=====================# 


#======================Dashboard session=====================# 
def generate_pdf(sales_data, start_date, end_date, filter_type):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Define Styles
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        'Header', fontSize=10, textColor=colors.black, spaceAfter=6, alignment=1
    )
    title_style = ParagraphStyle('Title', fontSize=14, spaceAfter=12, alignment=1)

    # Company Info Header
    company_info = Paragraph(
        "<b>Sound wave</b><br/>"
        "Email: soundwave01@gmail.com<br/>"
        "Website: www.soundwaveofficial.com", header_style
    )
    elements.append(company_info)
    elements.append(Spacer(1, 12))  # Space between header and title

    # PDF Title
    title = f"Sales Report ({start_date} - {end_date}) - Sorted by {filter_type.capitalize() if filter_type else 'Date'}"
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))  

    # Summary Data
    total_sales_price = Order.objects.aggregate(total_price_sum=Sum('total_price'))['total_price_sum']
    total_order = Order.objects.count()
    total_unit_sold = int(Order_items.objects.exclude(status='Cancelled').aggregate(Sum('quantity'))['quantity__sum'] or 0)
    total_discount_amount = Order.objects.filter(payment_status='Success').aggregate(total_discount=Sum(F('offer_price') + F('coupon_price')))['total_discount'] or 0.00
    cancelled_orders_count = Order_items.objects.filter(status='Cancelled').count()
    cancelled_orders_amount = Order_items.objects.filter(status='Cancelled').aggregate(Sum('subtotal_price'))['subtotal_price__sum'] or 0.00

    summary_data = [
        ["Total Sales Price", f"{total_sales_price:.2f}"],
        ["Total Orders", total_order],
        ["Total Units Sold", total_unit_sold],
        ["Total discount Amount",f"{total_discount_amount:.2f}"],
        ["Total Cancel Orders",f"{cancelled_orders_count}"],
        ["Total Cancel Orders Amount",f"{cancelled_orders_amount}"],

        
    ]
    summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 24))

    # Table Header and Data
    table_data = [
        ["Date", "Total Sales Revenue",  "Number of Orders", "Total Items Sold"]
    ]

    for record in sales_data:
        table_data.append([
            record.get('date', 'N/A').strftime('%Y-%m-%d') if record.get('date') else 'N/A',
            f"{record.get('total_sales', 0):.2f}",
            record.get('number_of_orders', 0),
            record.get('total_items_sold', 0)
        ])

    table = Table(table_data, colWidths=[1*inch] + [1.5*inch]*4)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="sales_report.pdf"'
    return response


def generate_excel(sales_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    ws['A1'] = "Summary"
    ws['A1'].font = Font(bold=True)

    total_sales_price = Order.objects.aggregate(total_price_sum=Sum('total_price'))['total_price_sum']
    total_order = Order.objects.count()
    total_unit_sold = int(Order_items.objects.exclude(status='Cancelled').aggregate(Sum('quantity'))['quantity__sum'] or 0)
    total_discount_amount = Order.objects.aggregate(total_discount=Sum(F('offer_price') + F('coupon_price')))['total_discount'] or 0.00

    summary = [
        ("Total Sales Price", total_sales_price),
        ("Total Orders", total_order),
        ("Total Units Sold", total_unit_sold),
        ("Total Discount Amount", total_discount_amount)
    ]
    for row_num, (label, value) in enumerate(summary, start=2):
        ws[f'A{row_num}'] = label
        ws[f'B{row_num}'] = value
        ws[f'A{row_num}'].font = Font(bold=True)

    # Column headers
    headers = ["Date", "Revenue", "Orders Count", "Items Sold"]
    column_widths = [20, 20, 18, 15, 15] 
    ws.append(headers)

    # Set column widths
    for i, width in enumerate(column_widths, start=1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = width
    
    for record in sales_data:
        ws.append([
            record['date'],
            record['total_sales'],
            record['number_of_orders'],
            record['total_items_sold']
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'
    wb.save(response)
    return response


@never_cache
@user_passes_test(lambda u: u.is_superuser, login_url='/adminn/')
def admin_home(request):
    today = timezone.now().date()

    filter_type = request.GET.get('filter_type', 'all')   
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if filter_type == 'day':
        start_date = today
        end_date = today
    elif filter_type == 'month':
        start_date = today.replace(day=1)
        end_date =today
    elif filter_type == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today
    elif filter_type == 'custom':
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else today  
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else today

    else:
        start_date = today - timedelta(days=30)
        end_date =today

    orders = Order_items.objects.filter(order__created_at__gte=start_date)
    total_users = CustomUser.objects.exclude(is_staff=True).count()
    # Filter sales data by date range
    sales_data = Order.objects.filter(created_at__gte=start_date, created_at__lte=end_date+timedelta(days=1)) \
                            .annotate(date=TruncDate('created_at')) \
                            .values('date') \
                            .annotate(total_sales=Sum('total_price'),
                                      number_of_orders=Count('order_id'),
                                      total_items_sold=Sum('order_items__quantity')
                            ).order_by('date')
    
    total_revenue=Order.objects.filter(payment_status='Success').aggregate(total=Sum('total_price'))['total'] or 0.00
    total_orders=Order.objects.count()
    overall_discount = Order.objects.filter(payment_status='Success').aggregate(total_discount=Sum(F('offer_price') + F('coupon_price')))['total_discount'] or 0.00

    top_selling_products = (
        Order_items.objects.exclude(status='Cancelled')
        .values('variant__product__id')
        .annotate(total_sold=Sum('quantity'))
        .order_by('-total_sold')[:5]
    )

    products = []
    for item in top_selling_products:
        product = Product.objects.get(id=item['variant__product__id'])
        variant = Variant.objects.filter(product=product).first()
        products.append({
            'name': product.name,
            'price': product.price,
            'total_sold': item['total_sold'],
            'image': variant.image1.url if variant and variant.image1 else None
        })


    top_selling_brands=(
        Order_items.objects.exclude(status='Cancelled').values('variant__product__brand__name')
        .annotate(total_sold=Sum('quantity'))
        .order_by('-total_sold')[:5]
    )

    top_selling_categories=(
        Order_items.objects.exclude(status='Cancelled')
        .values('variant__product__category__name')
        .annotate(total_sold=Sum('quantity'))
        .order_by('-total_sold')[:5]
    )

    
    # Week-wise Sales  Data
    week_sales_data = Order.objects.annotate(week=TruncWeek('created_at')) \
                                .values('week') \
                                .annotate(total_sales=Sum('total_price')) \
                                .order_by('week')

    # Month-wise Sales Data
    month_sales_data = Order.objects.annotate(month=TruncMonth('created_at')) \
                                    .values('month') \
                                    .annotate(total_sales=Sum('total_price')) \
                                    .order_by('month')

    # year-wise Sales Data
    year_sales_data = Order.objects.annotate(year=TruncYear('created_at')) \
                                .values('year') \
                                .annotate(total_sales=Sum('total_price')) \
                                .order_by('year')
    

    day_labels = [sale['date'].strftime('%Y-%m-%d') for sale in sales_data]  
    day_data = [float(sale['total_sales']) if sale['total_sales'] is not None else 0.0 for sale in sales_data] 

    week_labels = [sale['week'].strftime('%Y-%m-%d') for sale in week_sales_data]
    week_data = [float(sale['total_sales']) if sale['total_sales'] is not None else 0.0 for sale in week_sales_data]

    month_labels = [sale['month'].strftime('%Y-%m') for sale in month_sales_data]
    month_data = [float(sale['total_sales']) if sale['total_sales'] is not None else 0.0 for sale in month_sales_data]

    year_labels = [sale['year'].strftime('%Y') for sale in year_sales_data]
    year_data = [float(sale['total_sales']) if sale['total_sales'] is not None else 0.0 for sale in year_sales_data]

    # Handling PDF and Excel download
    if request.GET.get('download') == 'pdf':
        buffer = generate_pdf(sales_data, start_date, end_date,filter_type)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_{end_date}.pdf"'
        return response

    if request.GET.get('download') == 'excel':
        buffer = generate_excel(sales_data)
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_{end_date}.xlsx"'
        return response

    return render(request, 'admin/homepage.html', {
        'total_users': total_users,
        'sales_data': sales_data,
        'filter_type': filter_type,
        'start_date': start_date,
        'end_date': end_date,
        'total_revenue':total_revenue,
        'total_orders':total_orders,
        'overall_discount':overall_discount,
        'products': products,        
        'top_selling_brands':top_selling_brands,
        'top_selling_categories':top_selling_categories,
        'day_labels':json.dumps(day_labels),
        'day_data':json.dumps(day_data),
        'week_labels':json.dumps(week_labels),
        'week_data':json.dumps(week_data),
        'month_labels':json.dumps(month_labels),
        'month_data':json.dumps(month_data),
        'year_labels':json.dumps(year_labels),
        'year_data':json.dumps(year_data),    
    })

#======================Dashboard session End=====================# 


#======================User managment session=====================# 
@never_cache
@user_passes_test(lambda u : u.is_superuser,login_url='/adminn/')
def user_managment(request):
    user_info=CustomUser.objects.filter(is_staff=False)
    return render(request,'admin/userview.html',{'user_info':user_info})


@never_cache
@user_passes_test(lambda u : u.is_superuser,login_url='/adminn/')
def block_user(request,id):
    user=get_object_or_404(User,id=id)
    user.is_active=False
    user.save()
    return redirect('list_user')


@never_cache
@user_passes_test(lambda u : u.is_superuser,login_url='/adminn/')
def unblock_user(requset,id):
    user=get_object_or_404(User,id=id)
    user.is_active=True
    user.save()
    return redirect('list_user')

#======================User managment session End=====================# 

